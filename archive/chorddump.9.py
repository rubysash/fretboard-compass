"""
Takes a chord database, in xlsx and converts to various charts
scales, progressions, etc.

The idea is to show exactly what to play for both scales and progressions
for my favorite progressions (and all progressions), with a quick reference
for people like me who might know how to put fingers on a guitar, but
not know what to play together, or understand much about music theory, but
if given some chords and scales can have fun playing it.


requires 3 tabs on the input file:  chords, favs, scales.

chords format:
VERIFIED	KEY	VARIANT	NAME	POSITION	DOTS
1	X02220	0	A	1	[320,130,2],[340,130,1],[360,130,3]
1	X03211	0	A	5	[320,160,4],[340,130,2],[360,100,1],[380,100,1]

favs format:
VERIFIED	KEY	VARIANT	NAME	POSITION	DOTS
1	I II III iii D	0	D	1	D E C# F#m
1	I iii IV V Ballads C	0	C	1	C Em F G


scales format example:
VERIFIED	KEY	VARIANT	NAME	POSITION	DOTS
1	A	0	A	1	[0,2,4],[100,2,4],[0,2,4],[1,102,4],[0,2,3],[0,2,4]
1	B	0	B	1	[0,2,4],[1,102,4],[1,2,4],[1,3,104],[100,2,4],[0,2,4]


Examples of how to run the program depending on what you want to generate:

(chorddumper) u@h:~/chorddumper$ python3 chorddump.7.py -i chords7.xlsx -f
  Writing Favs:  7.favs.html
(chorddumper) u@h:~/chorddumper$ python3 chorddump.7.py -i chords7.xlsx -s
Writing Scales:  7.scales.html
(chorddumper) u@h:~/chorddumper$ python3 chorddump.7.py -i chords7.xlsx -c
Writing Chords:  7.all.html
(chorddumper) u@h:~/chorddumper$ python3 chorddump.7.py -i chords7.xlsx -a
  Writing Favs:  7.favs.html
Writing Chords:  7.all.html
Writing Scales:  7.scales.html

Features:
Prefers first position in database in the case of duplicate chords or voicings
If "Verified" is 0, it will highlight the chord/scale in red so you can verify it
If Verified is 1, it will print normal.
Dumps list of scales in Database (Currently just Major and Minor)
Dumps all chords in DB (Currently 102)
Dumps all favorite progressions (34 currently), combining scales with progressions
Attempts intelligent page breaks for easy printing

Todo:
DONE: Move svg and html into container folders instead of in the root
ONGOING: Add more data to the chord database (full scales, full progressions, etc)
Automate dot position base don key, (not sure how to show finger placements though)
Update database instead of using the lame offset hack
DONE: Does not allow duplicate fingerings of different chords (D#m/Eb,Abm/G#m etc)
DONE: Cannot print favorites until you first print out scales and chords, detect or force

Tested on:
Debian 12 and Windows 10

"""

# import base64
import collections  # to sort dictionary and maintain order
import glob  # for glob stuff that lists all files in a directory
import json  # to read/write json
import os  # for clear screen function and full module required for webbrowser + chrome / bug
import sys # for argv and sys.exit
import getopt  # for argv and sys.exit
import webbrowser  # to open automagically in browser

# Third-party imports, sorted alphabetically
from colorama import init, Fore, Back, Style  # colorize
from openpyxl import load_workbook  # for spreadsheet stuff
import svgwrite  # to create svg graphics

# Initialization of certain modules
init()  # required for colorama colors

VERSION = 9

# create files directory
output_dir = f"./{VERSION}"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    
# fixme: lame hack to use existing data as is.
XOFFSET = -260
YOFFSET = -40

# line cooridinates
ALLX = [280, 300, 320, 340, 360, 380]  # the x cooridinates
ALLY = [80, 110, 140, 170, 200]  # the y cooridinates

# screen creation for the scale/chord
SCREENX = 135  # x dim of image created
SCREENY = 180  # y dim of image created

# output files are html which load the svg
ALLFILE = os.path.join(output_dir, f"{VERSION}.chords.html")
FAVFILE = os.path.join(output_dir, f"{VERSION}.favs.html")
SCALESFILE = os.path.join(output_dir, f"{VERSION}.scales.html")


def get_header(title):
    """Generate an HTML header with a specified title.

    Args:
        title (str): The title to be used in the header.

    Returns:
        str: An HTML formatted header as a string.
    """
    header = (
        """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <meta http-equiv="X-UA-Compatible" content="ie=edge" />
    <title>""" + title + """</title>
    <style>
        table, th, td {
            border: 0px solid black;
        }
        body {
            font: bold italic 15px/20px arial,sans-serif;
            color: black;
        }
        td { width: 140px; height: 185px; }
        svg {
            position: absolute;
            width: 100%;
            height: 100%;
        }
        @media print {
            body, page {
                margin: 0;
                box-shadow: 0;
            }
            .no-break {
                page-break-inside: avoid;
            }
        }
    </style>
</head>
<body>"""
    )
    return header


def create_html(d, htmlfile, colmax, title):
    """Create an HTML file based on a dictionary of data, a file name, and a maximum column count.

    Args:
        d (dict): A dictionary containing chord or scale data.
        htmlfile (str): The name of the output HTML file.
        colmax (int): The maximum number of columns per row in the table.
    """
    # fixme:  use a tempalte
    header = get_header(title)

    # fixme: don't think I want to do this at all
    # this line will print all files in 1101001 pattern order
    # files = glob.glob("*." + filetype)

    count = 0
    with open(htmlfile, "w", encoding="utf-8") as html:

        html.write(header)
        html.write("<table><tr>")
        for f in d:  # if printing chord database we passed to it

            # print(d[f]['chord'])
            filename = f + ".svg"  # if printing chord database we passed
            count += 1
            # use modulos to determine how wide the rows are
            if count % colmax:
                html.write("<td><img src='" + filename + "' /></td>\n")
            else:
                html.write("<td><img src='" + filename + "' /></td>\n")
                html.write("</tr></table>\n<table><tr>\n")
        # finalize our html so it's valid
        html.write("</tr></table>\n</body></html>")

    open_html(htmlfile)


def open_html(htmlfile):
    """Open an HTML file in the default web browser."""
    filepath = os.path.join(os.getcwd(), htmlfile)  # Ensure correct path construction
    if os.name == "nt":
        webbrowser.get("C:/Program Files/Google/Chrome/Application/chrome.exe %s").open(
            "file://" + filepath
        )
    else:
        webbrowser.open("file://" + filepath)


def clear():
    """Clear the console screen."""
    # Use os.name to check the operating system type
    if os.name == "nt":  # If the operating system is Windows
        _ = os.system("cls")
    else:  # For Unix and Unix-like systems
        _ = os.system("clear")


def list_files(filetype):
    """Print a list of files of a specified filetype present in the current directory.

    Args:
        filetype (str): The file extension to search for.
    """
    files = glob.glob("*." + filetype)
    print(Fore.RED + "-" * 55)
    print(Fore.RED + "Compatible Files Found:")
    print(Fore.RED + "-" * 55)
    for f in files:
        print(Fore.GREEN + f, end="\t")
    print("\n")


def make_dots(x, y, r, color):
    """Create SVG dots at specified coordinates, radius, and color.

    Args:
        x (float): The x-coordinate of the dot's center.
        y (float): The y-coordinate of the dot's center.
        r (int): The radius of the dot.
        color (str): The fill color of the dot.
    """
    new_x = x + XOFFSET
    new_y = y + YOFFSET

    dwg.add(
        dwg.circle(
            center=(new_x, new_y), r=r, stroke=svgwrite.rgb(10, 10, 10), fill=color
        )
    )


def make_squares(x1, y1, x2, y2, width, fillcolor):
    """Create SVG rectangles with specified corner coordinates, width, and fill color.

    Args:
        x1 (float): The x-coordinate of the near corner.
        y1 (float): The y-coordinate of the near corner.
        x2 (float): The x-coordinate of the far corner.
        y2 (float): The y-coordinate of the far corner.
        width (int): The stroke width of the rectangle.
        fillcolor (str): The fill color of the rectangle.
    """
    dwg.add(
        dwg.rect(
            (x1, y1),
            (x2, y2),
            stroke=svgwrite.rgb(10, 10, 20),
            stroke_width=width,
            fill=fillcolor,
        )
    )


def make_lines(x1, y1, x2, y2, width):
    """Draw SVG lines between specified start and end coordinates with a certain width.

    Args:
        x1 (float): The starting x-coordinate.
        y1 (float): The starting y-coordinate.
        x2 (float): The ending x-coordinate.
        y2 (float): The ending y-coordinate.
        width (int): The stroke width of the line.
    """
    x1 += XOFFSET
    y1 += YOFFSET
    x2 += XOFFSET
    y2 += YOFFSET

    dwg.add(
        dwg.line(
            start=(x1, y1),
            end=(x2, y2),
            stroke=svgwrite.rgb(10, 10, 20),
            stroke_width=width,
        )
    )


def make_labels(x, y, size, alpha):
    """Create SVG text elements at specified coordinates with a certain size.

    Args:
        x (float): The x-coordinate where the text will be inserted.
        y (float): The y-coordinate where the text will be inserted.
        size (str): The font size of the text.
        alpha (str): The text content.
    """
    x += XOFFSET
    y += YOFFSET
    dwg.add(
        dwg.text(
            alpha,
            insert=(x, y),
            stroke="none",
            fill=svgwrite.rgb(15, 15, 15),
            # does not work with tiny profile
            font_size=size,
            font_weight="bold",
            style="font-family:consolas",
        )
    )


def fix_list(dots):
    """Convert a string representation of dot positions to a list.

    Args:
        dots (str): A string representation of dot positions.

    Returns:
        list: A list of dot positions extracted from the string.
    """
    crap = []
    xys = dots.split("],[")
    for a in xys:
        a = a.replace("[", "")
        a = a.replace("]", "")
        crap.append(a)
    return crap


def get_raw(sheet):
    data = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=8):  # Assuming the "VARIANT" column is the 5th column
        verified = row[0].value
        key = row[1].value
        variant = row[2].value  # This is the new "VARIANT" column
        name = row[3].value
        position = row[4].value
        dots = row[5].value

        if key is None:
            break
        
        # The unique key now includes the "VARIANT"
        unique_key = f"{key}_{variant}_{position}"
        data[unique_key] = {
            "verified": verified,
            "key": key,
            "variant": variant,
            "name": name,
            "position": position,
            "dots": dots,
        }

    return collections.OrderedDict(sorted(data.items()))


def get_chord_variant(chord_name, chords):
    # Find all variants for the given chord name
    variants = [
        chord['variant'] for chord in chords.values() if chord['name'] == chord_name
    ]
    # Convert variants to integers to sort them, then return the lowest
    variants = list(map(int, variants))
    return str(min(variants)) if variants else '0'


def dump_favs(ffile, favorites, scales, chords):
    header = get_header("Just the Favs")

    print("  Writing Favs: ", ffile)
    with open(ffile, "w", encoding="utf-8") as favs:
        favs.write(header)

        for fav_key in favorites:
            fav_name = fav_key.split("_", maxsplit=1)[0]
            favs.write(f"<div class='no-break'><h3>{fav_name}</h3>\n<table>\n")
            
            # fixme: update logic for scale variant 
            scale_variant = '0'
            # Get chord variant from chords data
            chord_variant = get_chord_variant(favorites[fav_key]['name'], chords)
            
            dot = favorites[fav_key]["dots"]
            key_name = favorites[fav_key]["name"]

            # Retrieve the scale SVG if it exists
            scale_key = f"{key_name}_{scale_variant}_{favorites[fav_key]['position']}"
            if scale_key in scales:
                scalesvg = f"{scales[scale_key]['key']}_{scale_variant}_{scales[scale_key]['position']}.svg"
                favs.write(f"<tr><td><img src='{scalesvg}' alt='{key_name} scale' /></td>\n")

            dots = dot.split(" ")
            columns = 0
            for dot in dots:
                # Get all chords that match the name
                matching_chords = [chord for chord in chords.values() if chord["name"] == dot]
                # Sort them by position, preferring '1' if it exists
                matching_chords.sort(key=lambda c: (c["position"] != "1", c["position"]))
                if matching_chords:
                    chord_key = matching_chords[0]["key"]
                    chord_position = matching_chords[0]["position"]
                    chord_svg = f"{chord_key}_{chord_variant}_{chord_position}.svg"
                    if columns == 4:
                        favs.write("</tr>\n<tr>")  # End the current row and start a new one
                        columns = 0
                    favs.write(f"<td><img src='{chord_svg}' alt='{dot}' /></td>\n")
                    columns += 1

            favs.write("</tr>\n</table>\n</div>\n\n")  # Close the last row and the table

        favs.write("</body></html>")

    open_html(ffile)

    
def dump_scales(scales):
    """Create SVG files for all scales in the provided dictionary.

    Args:
        scales (dict): A dictionary of scales to process.
    """
    s_data = {}
    for key in scales:

        # extract chunks into friendly names
        verified = scales[key]["verified"]
        name = scales[key]["name"]
        position = scales[key]["position"]
        dots = scales[key]["dots"]

        # build this dictionary "value"; None is allowed, in case something is missing
        thisdict = {"name": name, "data": dots}

        # work around the sheet.max_row bug thing
        if key is None:
            break
        else:
            # build unique key and load thisdict into it
            key = str(key)
            s_data[key] = thisdict

            # and the filename
            svgout = os.path.join(output_dir, f"{key}.svg")

            # fixme:  is there a better way than global?
            global dwg
            dwg = svgwrite.Drawing(
                svgout,
                (int(SCREENX), int(SCREENY)),
                viewBox=(0, 0, int(SCREENX), int(SCREENY)),
                fill="#FF0000",  # fixme: color isn't shown, for debugging, remove
                # profile='tiny' # breaks font/style on labels
            )

            # just makes a nice border around each picture
            if str(verified) == str(0):
                make_squares(
                    0, 0, SCREENX, SCREENY, 1, "pink"
                )  # the border # fixme: pink
            else:
                make_squares(0, 0, SCREENX, SCREENY, 1, "#EEEEEE")  # the border

            for y in ALLY:  # draw our fret lines
                make_lines(ALLX[5], y, ALLX[0], y, 3)

            for x in ALLX:  # draw our string lines
                make_lines(x, ALLY[0], x, ALLY[4], 1)

            make_labels(261, 89, 16, position)  # which fret this pattern starts on
            offset = 330 - len(name) * 10 / 2  # a crude attempt at centering label
            make_labels(offset, 58, 18, name)  # the actual chordname at top

            # this is a work around because I loaded from sheet instead of json
            # converts [x,y,z],[x,y,z] into a true list from a string
            # fixme:  real data probably coming as json from sqlite
            fixedlist = fix_list(dots)
            # "EMaj_1": {
            #    "name": "EMaj",
            #    "data": "[0,2,4],[0,2,4],[1,2,4],[1,2,4],[0,2,4],[0,2,4]"
            # },

            if len(fixedlist) > 1:

                # however many dots we have, loop and place them
                for i in range(len(fixedlist)):

                    # chunk is 3 parts, we are moving around as needed to display them
                    chunk = fixedlist[i]
                    fret_pos = chunk.split(",")

                    # pdb.set_trace()
                    for text_y in fret_pos:
                        y = int(text_y)
                        # i = one [] of data, which coincides with a string
                        xx = float(ALLX[int(i)])
                        # y = one of the fret positions (0,2,3)
                        if y >= 100:
                            y += -100
                            # fixme:  these do not scale, fix the offsets and radius to scale
                            yy = float(ALLY[y]) - 10 # Raise it a bit so it's above the fret
                            make_dots(
                                xx, yy, 8, "black"
                            )  # this is the finger position, the actual dot
                        else:
                            yy = float(ALLY[y]) - 10 # Raise it a bit so it's above the fret
                            make_dots(xx, yy, 8, "white")
            else:
                pass

            # write the file
            dwg.save()

    d_sorted = collections.OrderedDict(
        sorted(s_data.items(), key=lambda t: t[1]["name"])
    )
    print("Writing Scales: ", SCALESFILE)
    create_html(d_sorted, SCALESFILE, 4, "All Scales in DB")


def get_data(inputfile, data):
    """Retrieve data from an Excel file for a specified sheet.

    Args:
        inputfile (str): The path to the Excel file.
        data (str): The name of the sheet to retrieve data from.

    Returns:
        dict: A dictionary containing the data from the specified sheet.
    """
    try:
        workbook = load_workbook(filename=inputfile, data_only=True)

        data_set = get_raw(workbook[data]) if data in workbook else None

        if data is None:
            missing_sheets = [sheet for sheet in [data] if sheet not in workbook]
            print(
                f"Error: Could not find the following sheet(s): {', '.join(missing_sheets)}"
            )
            return

        return data_set

    except Exception as e:
        print(f"An error occurred: {e}")
        return


def dump_all(inputfile):
    """Generate all possible outputs (favorites, chords, scales) from an Excel file.

    Args:
        inputfile (str): The path to the Excel file.
    """
    try:
        workbook = load_workbook(filename=inputfile, data_only=True)
        chords_sheet = "chords"
        favs_sheet = "favs"
        scales_sheet = "scales"

        chords = get_raw(workbook[chords_sheet]) if chords_sheet in workbook else None
        favs = get_raw(workbook[favs_sheet]) if favs_sheet in workbook else None
        scales = get_raw(workbook[scales_sheet]) if scales_sheet in workbook else None

        if chords is None or favs is None or scales is None:
            missing_sheets = [
                sheet
                for sheet in [chords_sheet, favs_sheet, scales_sheet]
                if sheet not in workbook
            ]
            print(
                f"Error: Could not find the following sheet(s): {', '.join(missing_sheets)}"
            )
            return

        # chords and scales first
        dump_chords(chords)
        dump_scales(scales)
        dump_favs(FAVFILE, favs, scales, chords)
        
    except Exception as e:
        print(f"An error occurred: {e}")
        return


def dump_json(inputfile):
    """Generate all possible outputs (favorites, chords, scales) from an Excel file.

    Args:
        inputfile (str): The path to the Excel file.
    """
    try:
        workbook = load_workbook(filename=inputfile, data_only=True)
        chords_sheet = "chords"
        favs_sheet = "favs"
        scales_sheet = "scales"

        chords = get_raw(workbook[chords_sheet]) if chords_sheet in workbook else None
        favs = get_raw(workbook[favs_sheet]) if favs_sheet in workbook else None
        scales = get_raw(workbook[scales_sheet]) if scales_sheet in workbook else None

        if chords is None or favs is None or scales is None:
            missing_sheets = [
                sheet
                for sheet in [chords_sheet, favs_sheet, scales_sheet]
                if sheet not in workbook
            ]
            print(
                f"Error: Could not find the following sheet(s): {', '.join(missing_sheets)}"
            )
            return

        print(Fore.RED + '-'*55)
        print(Fore.RED + "CHORDS:")
        print(Fore.RED + '-'*55 + Fore.GREEN)
        print(Style.RESET_ALL)
        print(json.dumps(chords, indent=2))

        print(Fore.RED + '-'*55)
        print(Fore.RED + "SCALES:")
        print(Fore.RED + '-'*55 + Fore.GREEN)
        print(Style.RESET_ALL)
        print(json.dumps(scales, indent=2))

        print(Fore.RED + '-'*55)
        print(Fore.RED + "FAVORITES:")
        print(Fore.RED + '-'*55 + Fore.GREEN)
        print(Style.RESET_ALL)
        print(json.dumps(favs, indent=2))

        
    except Exception as e:
        print(f"An error occurred: {e}")
        return

    
def dump_chords(chords):
    """Create SVG files for all chords in the provided dictionary.

    Args:
        chords (dict): A dictionary of chords to process.
    """
    c_data = {}
    for key in chords:
        # extract chunks into friendly names
        verified = chords[key]["verified"]
        name = chords[key]["name"]
        position = chords[key]["position"]
        dots = chords[key]["dots"]

        # build this dictionary "value"; None is allowed, in case something is missing
        thisdict = {"name": name, "data": dots}

        # work around the sheet.max_row bug thing
        if key is None:
            break
        else:
            # build unique key and load thisdict into it
            c_data[key] = thisdict

            # and the filename
            svgout = os.path.join(output_dir, f"{key}.svg")

            # fixme:  is there a better way than global?
            global dwg
            dwg = svgwrite.Drawing(
                svgout,
                (int(SCREENX), int(SCREENY)),
                viewBox=(0, 0, int(SCREENX), int(SCREENY)),
                fill="#FF0000",  # fixme: color isn't shown, for debugging, remove
                # profile='tiny' # breaks font/style on labels
            )

            # just makes a nice border around each picture
            if str(verified) == str(0):
                make_squares(
                    0, 0, SCREENX, SCREENY, 1, "pink"
                )  # the border # fixme: pink
            else:
                make_squares(0, 0, SCREENX, SCREENY, 1, "#EEEEEE")  # the border

            # print the x and o at top
            # fixme:  data input could be corrupt, validate it x X o 0 O
            patterncount = 0
            for x in key:
                if x == "_":  # was failing if key had zero in position XXXXXX_10
                    break
                if x == "X":
                    make_labels(ALLX[patterncount] - 4, 75, 14, "X")
                if x == "0":
                    make_labels(ALLX[patterncount] - 4, 75, 14, "O")
                patterncount += 1

                # find the typo fixme:  single use to search my db for typo
                # if x =='x':
                #    print(name,key,position)

            for y in ALLY:  # draw our fret lines
                make_lines(ALLX[5], y, ALLX[0], y, 3)

            for x in ALLX:  # draw our string lines
                make_lines(x, ALLY[0], x, ALLY[4], 1)

            make_labels(261, 89, 16, position)  # which fret this pattern starts on
            offset = 330 - len(name) * 10 / 2  # a crude attempt at centering label
            make_labels(offset, 58, 18, name)  # the actual chordname at top

            # this is a work around because I loaded from sheet instead of json
            # converts [x,y,z],[x,y,z] into a true list from a string
            # fixme:  real data probably coming as json from sqlite
            fixedlist = fix_list(dots)

            if len(fixedlist) >= 1:

                # however many dots we have, loop and place them
                for i in range(len(fixedlist)):

                    # chunk is 3 parts, we are moving around as needed to display them
                    # xyz[0] is x
                    # xyz[1] is y
                    # xyz[2] is the finger to use
                    chunk = fixedlist[i]
                    xyz = chunk.split(",")

                    xx = float(xyz[0])
                    yy = float(xyz[1])
                    text = 0  # give zz a default

                    if len(xyz) > 2:  # but if zz exists, then use it
                        text = xyz[2]

                    # fixme:  these do not scale, fix the offsets and radius to scale
                    make_dots(
                        xx, yy, 9, "white"
                    )  # this is the finger position, the actual dot
                    make_labels(
                        xx - 4, yy + 4, "14px", text
                    )  # this is the finger number inside the dot
            else:
                pass

            # write the file
            dwg.save()

    # requires "collections" module, sorts our dicitonary
    d_sorted = collections.OrderedDict(
        sorted(c_data.items(), key=lambda t: t[1]["name"])
    )

    print("Writing Chords: ", ALLFILE)
    create_html(d_sorted, ALLFILE, 5, "All Chords in DB")


def help_message(msg):
    """Display a help message with instructions and available commands.

    Args:
        msg (str): The message to be displayed.
    """
    clear()
    print(Fore.RED + "-" * 55)
    print(Fore.RED + msg)
    print(Fore.RED + "-" * 55 + Fore.GREEN)
    # print(Style.BRIGHT + Fore.YELLOW + "\n(Argument Order is Specific!!)\n")
    print(Style.RESET_ALL + "Examples:")
    print(sys.argv[0] + " -h")
    #print(sys.argv[0] + " -i <inputfile> -f")
    print(sys.argv[0] + " -i <inputfile> -j")
    print(sys.argv[0] + " -i <inputfile> -c")
    print(sys.argv[0] + " -i <inputfile> -f")
    print(sys.argv[0] + " -i <inputfile> -s")
    print(sys.argv[0] + " -i <inputfile> -a")

    print(Style.BRIGHT + Fore.YELLOW + "FLAG EXPLANATIONS\n")
    print("-h prints this help message")
    print("-i (INPUTFILE) inputfile is your source, input file, this is required")
    print("-f (FAVORITES) prints my favorite progressions")
    print("-j (JSON) dumps json to screen")
    print("-a (ALL CHORDS, SCALES, FAVS) svg files and loads them into chrome to view")
    print("-s (SCALES) creates all known scales files")
    print("-c (CHORDS) creates all known chords files")
    print(Style.RESET_ALL)
    list_files("xlsx")
    sys.exit(2)


def is_accessible(path, mode="r", encoding="utf-8"):
    """Check if a file or directory at a given path is accessible.

    Args:
        path (str): The file or directory path to check.
        mode (str): The mode in which the file should be accessible.
        encoding (str): The encoding to use (relevant for text modes).

    Returns:
        bool: True if the resource is accessible, False otherwise.
    """
    try:
        # Only include encoding if mode indicates a text file
        if "b" not in mode:
            f = open(path, mode, encoding=encoding)
        else:
            f = open(path, mode)
        f.close()
    except IOError:
        return False
    return True


def main():
    """Main function that parses command line arguments and runs the appropriate processes."""
    inputfile = ""
    dump_options = {"-j": False, "-c": False, "-s": False, "-a": False}

    try:
        opts, args = getopt.getopt(sys.argv[1:], "h?ajcfsi:", ["help", "?"])
    except getopt.GetoptError as err:
        help_message("BAD OPTION")

    for opt, arg in opts:
        if opt in ("-h", "--help", "-?"):
            help_message("INSTRUCTIONS")
        if opt in ("-i",):
            if is_accessible(arg):
                inputfile = arg
            else:
                help_message("CANNOT READ FILE")
        if opt in dump_options.keys():
            dump_options[opt] = True

    if not inputfile:
        help_message("Input file not specified or not accessible.")

    if dump_options["-a"]:
        dump_all(inputfile)
    else:
        executed_any = False
        if dump_options["-j"]:  # Assuming '-j' is for JSON dump, though not mentioned
            dump_json(inputfile)
            executed_any = True
        if dump_options["-c"]:
            chords = get_data(inputfile, "chords")
            dump_chords(chords)
            executed_any = True
        if dump_options["-s"]:
            scales = get_data(inputfile, "scales")
            dump_scales(scales)
            executed_any = True

        if not executed_any:  # No valid dump option found
            help_message("No valid options provided for actions. Showing help.")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main()
    else:
        help_message("INSTRUCTIONS")

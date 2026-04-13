# Standard library imports, sorted alphabetically
#import base64
import collections  # to sort dictionary and maintain order
import glob  # for glob stuff that lists all files in a directory
import json  # to read/write json
import os  # for clear screen function and full module required for webbrowser + chrome / bug
import sys  # for argv and sys.exit
import webbrowser  # to open automagically in browser

# Third-party imports, sorted alphabetically
from colorama import init, Fore, Back, Style  # colorize
from openpyxl import load_workbook  # for spreadsheet stuff
import svgwrite  # to create svg graphics

# Initialization of certain modules
init()  # required for colorama colors


XOFFSET = (
    -260
)  # because I already had data in the wrong spot so I just globally moved it
YOFFSET = (
    -40
)  # because I already had data in the wrong spot so I just globally moved it

ALLX = [280, 300, 320, 340, 360, 380]  # the x cooridinates
ALLY = [80, 110, 140, 170, 200]  # the y cooridinates

SCREENX = 135  # x dim of image created
SCREENY = 180  # y dim of image created

ALLFILE = "5.all.html"
FAVFILE = "5.favs.html"
SCALESFILE = "5.scales.html"

'''
# https://stackoverflow.com/questions/22162321/search-for-a-value-in-a-nested-dictionary-python
def get_path(nested_dict, value, prepath=()):
    for k, v in nested_dict.items():
        path = prepath + (k,)
        if v == value:  # found value
            return path
        elif hasattr(v, "items"):  # v is a dict
            p = get_path(v, value, path)  # recursive call
            if p is not None:
                return p
        else
            pass
'''

def get_header(title):
    header = (
        """<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <meta http-equiv="X-UA-Compatible" content="ie=edge" />
    <title>"""
        + title
        + """</title>
  </head>
<style>
table, th, td {
  border: 0px solid black;
}
body{
    font: bold italic 15px/20px arial,sans-serif;
    color: black;
}
td { width: 140px ; height: 185px ; }
svg {
  position: absolute;
  width: 100%;
  height: 100%;
}
</style><body>
"""
    )
    return header


def create_html(d, htmlfile):
    # print(type(d))
    # sys.exit()
    """
    List Files that match
    """
    # generate a valid html header with basic styling
    # fixme:  templates?
    header = get_header("all chords in db")

    # fixme: don't think I want to do this at all
    # this line will print all files in 1101001 pattern order
    # files = glob.glob("*." + filetype)

    count = 0
    colmax = 6  # how many graphics per line
    with open(htmlfile, "w", encoding="utf-8") as chords:

        chords.write(header)
        chords.write("<table><tr>")
        for f in d:  # if printing chord database we passed to it

            # print(d[f]['chord'])
            filename = f + ".svg"  # if printing chord database we passed
            count += 1
            # use modulos to determine how wide the rows are
            if count % colmax:
                chords.write("<td><img src='" + filename + "' /></td>\n")
            else:
                chords.write("<td><img src='" + filename + "' /></td>\n")
                chords.write("</tr></table>\n<table><tr>\n")
        # finalize our html so it's valid
        chords.write("</tr></table>\n</body></html>")

    open_html(htmlfile)


def open_html(htmlfile):
    if os.name == "nt":
        webbrowser.get("C:/Program Files/Google/Chrome/Application/chrome.exe %s").open(
            "file://" + os.path.realpath(htmlfile)
        )
    else:
        webbrowser.open("file://" + os.path.realpath(htmlfile))


def clear():
    """
    Python doesn't have a "clear screen" function
    So we are making an OS friendly one
    requires os module
    """
    # Use os.name to check the operating system type
    if os.name == "nt":  # If the operating system is Windows
        _ = os.system("cls")
    else:  # For Unix and Unix-like systems
        _ = os.system("clear")


def list_files(filetype):
    """
    List Files that match whatever filetype name (not magic mime type)
    """
    files = glob.glob("*." + filetype)
    print(Fore.RED + "-" * 55)
    print(Fore.RED + "Compatible Files Found:")
    print(Fore.RED + "-" * 55)
    for f in files:
        print(Fore.GREEN + f, end="\t")
    print("\n")


def get_col(sheet, col):
    """
    dump the enter column we are looking for
    """
    print(Fore.RED + "-" * 55)
    print(Fore.RED + "DUMPING COL " + str(col))
    print(Fore.RED + "-" * 55 + Style.RESET_ALL)
    rowcount = 0
    # fixme:  considr adding max limit like 100 or so
    for row in sheet.iter_rows(min_row=2, values_only=True):
        this = row[col]
        if this is None:
            rowcount += 1
            pass
        else:
            rowcount += 1
            print(rowcount, row[col])


def count_cols(sheet):
    """
    Given a sheet, tell me the columns that have stuff
    Immediately returns at first blank header
    https://bitbucket.org/openpyxl/openpyxl/issues/514/cell-max_row-reports-higher-than-actual
    """
    max_col = 0
    for i in range(1, 200):
        # print(sheet.cell(row=2, column=i).value)
        if sheet.cell(row=2, column=i).value is None:
            max_col = i
            break
    return max_col


def count_rows(sheet):
    """
    Given a sheet, tell me the rows that have stuff
    Immediately returns at first blank
    https://bitbucket.org/openpyxl/openpyxl/issues/514/cell-max_row-reports-higher-than-actual
    """
    for i in range(1, 200):
        if sheet.cell(row=i, column=2).value is None:
            max_row = i
            break
    return max_row


def make_dots(x, y, r, color):
    """
    x,y is the center
    r is radius
    color is the fill color, stroke is default to black
    """
    new_x = x + XOFFSET
    new_y = y + YOFFSET

    dwg.add(
        dwg.circle(
            center=(new_x, new_y), r=r, stroke=svgwrite.rgb(10, 10, 10), fill=color
        )
    )


def make_squares(x1, y1, x2, y2, width, fillcolor):
    """
    makesquares:
    takes xy near and far corners
    draws rectangle using those points
    """
    dwg.add(
        dwg.rect(
            (x1, y1),
            (x2, y2),
            stroke=svgwrite.rgb(10, 10, 20),
            stroke_width=width,
            fill=fillcolor,
        )
    )


def make_lines(x1, y1, x2, y2, width):
    """
    takes xy start and stopcoords
    makes a line from xy1 to xy2
    """
    x1 += XOFFSET
    y1 += YOFFSET
    x2 += XOFFSET
    y2 += YOFFSET

    dwg.add(
        dwg.line(
            start=(x1, y1),
            end=(x2, y2),
            stroke=svgwrite.rgb(10, 10, 20),
            stroke_width=width,
        )
    )


def make_labels(x, y, size, alpha):
    """
    x,y start, font size in '12px' format, and then your text
    puts text on the screen.
    """
    x += XOFFSET
    y += YOFFSET
    dwg.add(
        dwg.text(
            alpha,
            insert=(x, y),
            stroke="none",
            fill=svgwrite.rgb(15, 15, 15),
            # does not work with tiny profile
            font_size=size,
            font_weight="bold",
            style="font-family:consolas",
        )
    )


def fix_list(dots):
    """
    I'm reading from xlsx so my cell.value doesn't look like a list
    this converts it into a list.  Hack so I can use my existing data
    """
    crap = []
    xys = dots.split("],[")
    for a in xys:
        a = a.replace("[", "")
        a = a.replace("]", "")
        crap.append(a)
    return crap


def get_raw(sheet):
    data = {}
    # skip header, sheet.max_row is busted
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=7):
        # extract chunks into friendly names
        verified = row[0].value
        key = row[1].value
        name = row[2].value
        position = row[3].value
        dots = row[4].value

        thisdict = {
            "verified": verified,
            "key": key,
            "name": name,
            "position": position,
            "dots": dots,
        }

        if key is None:
            break

        # build unique key and load thisdict into it
        key = str(key) + "_" + str(position)
        data[key] = thisdict

    # print("\n")
    # requires "collections" module, sorts our dicitonary
    # return collections.OrderedDict(sorted(data.items(), key=lambda t:t[1]["key"]))
    return data


def dump_json(sheet, cols):
    # skip header, sheet.max_row is busted, only want first 4 cols
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=cols):
        # extract chunks into friendly names
        #verified = row[0].value
        key = row[1].value
        name = row[2].value
        #position = row[3].value
        dots = row[4].value

        thisdict = {"name": name, "data": dots}

        if key is None:
            break

        # build unique key and load thisdict into it
        # key = str(pattern)+"_"+str(position)
        key = str(key)
        data[key] = thisdict

    # requires "collections" module, sorts our dicitonary
    d_sorted = collections.OrderedDict(sorted(data.items(), key=lambda t: t[1]["name"]))

    # sorting a dictionary of dictionaries w/o collections
    # for k, v in sorted(data.items(), key=lambda e: e[1]['chord']):
    #    print(k,v)

    print(Fore.RED + "-" * 55)
    print(Fore.RED + "SORTED JSON DUMP:")
    print(Fore.RED + "-" * 55 + Fore.GREEN)
    print(Style.RESET_ALL)
    print(json.dumps(d_sorted, indent=2))


def dump_header(sheet, cols):
    """
    Given a sheet, and colums, tell me the header names and first row values
    Uses count_cols and doesn't do more than it returns
    """
    # build our data first so we can format it together later
    row1 = []
    row2 = []

    # get the header
    for row in sheet.iter_rows(min_row=1, max_col=cols, max_row=1):
        for cell in row:
            row1.append(cell.value)

    # get first row
    for row in sheet.iter_rows(min_row=2, max_col=cols, max_row=2):
        for cell in row:
            row2.append(cell.value)

    # print header
    print(Fore.RED + "-" * 55)
    print(Fore.RED + f"{'#':>4}", end="\t")
    print(Fore.RED + f"{'HEADER':>22}", end="\t")
    print(Fore.RED + f"{'DATA'}")
    print(Fore.RED + "-" * 55)

    # print them out left and right justified and numbered for first row
    for i in range(0, cols):
        row1data = str(row1[i])
        row2data = str(row2[i])
        print(Fore.YELLOW + f"{i:>4}", end="\t")
        print(Fore.GREEN + f"{row1data:>22}", end="\t")
        print(Fore.YELLOW + f"{row2data}")


def dump_favs(ffile, favorites, scales, chords):
    """
      "i VI III VII Em": {
        "verified": 0,
        "key": "i VI III VII Em",
        "name": "Em",
        "position": 1,
        "dots": "Em CMaj GMaj DMaj"
      }

    "XXX442": {
        "verified": 1,
        "key": "XXX442",
        "name": "BMaj",
        "position": "1",
        "dots": "[380,130,1],[340,190,3],[360,190,4]"
    }

    "EMaj_1": {
        "verified": 1,
        "key": "EMaj_1",
        "name": "EMaj",
        "position": 1,
        "dots": "[0,2,4],[0,2,4],[1,2,4],[1,2,4],[0,2,4],[0,2,4]"
    },
    """

    header = get_header("Just the Favs")

    with open(ffile, "w", encoding="utf-8")  as favs:

        favs.write(header)

        for k in favorites:
            found = {}
            founds = []
            good_name = k.split("_")
            founds.append(f"<h3>{good_name[0]}</h3>\n<table><tr>\n")
            dot = favorites[k]["dots"]
            key_name = favorites[k]["name"]
            #thissvg = ""
            for s in scales:
                if key_name == scales[s]["name"]:
                    scalesvg = key_name + "_" + str(scales[s]["position"]) + ".svg"
                    founds.append("<td><img src='" + scalesvg + "' /></td>\n")

            dots = dot.split(" ")
            for dot in dots:
                for c in chords:
                    c_name = chords[c]["name"]
                    if c_name in found:
                        pass
                    else:
                        if dot == c_name:
                            found[c_name] = 1
                            founds.append("<td><img src='" + c + ".svg" + "' /></td>\n")

            for foundit in founds:
                favs.write(foundit)
            favs.write("</tr></table>\n\n")

        # finalize our html so it's valid
        favs.write("</tr></table>\n</body></html>")

    open_html(ffile)


def dump_scales(scales):
    s_data = {}
    for key in scales:

        # extract chunks into friendly names
        verified = scales[key]["verified"]
        name = scales[key]["name"]
        position = scales[key]["position"]
        dots = scales[key]["dots"]

        # build this dictionary "value"; None is allowed, in case something is missing
        thisdict = {"name": name, "data": dots}

        # work around the sheet.max_row bug thing
        if key is None:
            break
        else:
            # build unique key and load thisdict into it
            key = str(key)
            s_data[key] = thisdict

            # and the filename
            svgout = key + ".svg"

            # fixme:  is there a better way than global?
            global dwg
            dwg = svgwrite.Drawing(
                svgout,
                (int(SCREENX), int(SCREENY)),
                viewBox=(0, 0, int(SCREENX), int(SCREENY)),
                fill="#FF0000",  # fixme: color isn't shown, for debugging, remove
                # profile='tiny' # breaks font/style on labels
            )

            # just makes a nice border around each picture
            if str(verified) == str(0):
                make_squares(
                    0, 0, SCREENX, SCREENY, 1, "pink"
                )  # the border # fixme: pink
            else:
                make_squares(0, 0, SCREENX, SCREENY, 1, "#EEEEEE")  # the border

            # print the x and o at top
            # fixme:  data input could be corrupt, validate it x X o 0 O
            patterncount = 0
            for x in key:
                if x == "_":  # was failing if key had zero in position XXXXXX_10
                    break
                if x == "X":
                    make_labels(ALLX[patterncount] - 4, 75, 14, "X")
                if x == "0":
                    make_labels(ALLX[patterncount] - 4, 75, 14, "O")
                patterncount += 1

                # find the typo fixme:  single use to search my db for typo
                # if x =='x':
                #    print(name,key,position)

            for y in ALLY:  # draw our fret lines
                make_lines(ALLX[5], y, ALLX[0], y, 3)

            for x in ALLX:  # draw our string lines
                make_lines(x, ALLY[0], x, ALLY[4], 1)

            make_labels(261, 89, 16, position)  # which fret this pattern starts on
            offset = 330 - len(name) * 10 / 2  # a crude attempt at centering label
            make_labels(offset, 58, 18, name)  # the actual chordname at top

            # this is a work around because I loaded from sheet instead of json
            # converts [x,y,z],[x,y,z] into a true list from a string
            # fixme:  real data probably coming as json from sqlite
            fixedlist = fix_list(dots)

            # "EMaj_1": {
            #    "name": "EMaj",
            #    "data": "[0,2,4],[0,2,4],[1,2,4],[1,2,4],[0,2,4],[0,2,4]"
            # },

            if len(fixedlist) > 1:

                # however many dots we have, loop and place them
                for i in range(len(fixedlist)):

                    # chunk is 3 parts, we are moving around as needed to display them
                    chunk = fixedlist[i]
                    fret_pos = chunk.split(",")

                    for y in fret_pos:

                        # i = one [] of data, which coincides with a string
                        xx = float(ALLX[int(i)])
                        # y = one of the fret positions (0,2,3)
                        yy = float(ALLY[int(y)] - 10)
                        if y > 100:
                            y = -100
                            # fixme:  these do not scale, fix the offsets and radius to scale
                            make_dots(
                                xx, yy, 8, "black"
                            )  # this is the finger position, the actual dot
                        else:
                            make_dots(xx, yy, 8, "white")
            else:
                pass

            # write the file
            dwg.save()

    d_sorted = collections.OrderedDict(
        sorted(s_data.items(), key=lambda t: t[1]["name"])
    )
    create_html(d_sorted, SCALESFILE)


def dump_all(inputfile):
    """
    loops over spreadsheet
    builds json dictionary (original purpose of script)
    my tk.frets script uses json, but editing db was easier in spreadsheet..
    dumps to svg as it goes (was not much extra work to just do the task I wanted)
    """

    workbook = load_workbook(filename=inputfile, data_only=True)

    chords = get_raw(workbook["chords"])
    favs = get_raw(workbook["favs"])
    scales = get_raw(workbook["scales"])

    # print(Fore.RED + '-'*55)
    # print(Fore.RED + "SORTED JSON DUMP:")
    # print(Fore.RED + '-'*55 + Fore.GREEN)
    # print(Style.RESET_ALL)
    # print(json.dumps(chords, indent=2))
    # sys.exit()

    dump_favs(FAVFILE, favs, scales, chords)
    dump_chords(chords)
    dump_scales(scales)


def dump_chords(chords):
    c_data = {}
    for key in chords:

        # extract chunks into friendly names
        verified = chords[key]["verified"]
        name = chords[key]["name"]
        position = chords[key]["position"]
        dots = chords[key]["dots"]

        # build this dictionary "value"; None is allowed, in case something is missing
        thisdict = {"name": name, "data": dots}

        # work around the sheet.max_row bug thing
        if key is None:
            break
        else:
            # build unique key and load thisdict into it
            c_data[key] = thisdict

            # and the filename
            svgout = key + ".svg"

            # fixme:  is there a better way than global?
            global dwg
            dwg = svgwrite.Drawing(
                svgout,
                (int(SCREENX), int(SCREENY)),
                viewBox=(0, 0, int(SCREENX), int(SCREENY)),
                fill="#FF0000",  # fixme: color isn't shown, for debugging, remove
                # profile='tiny' # breaks font/style on labels
            )

            # just makes a nice border around each picture
            if str(verified) == str(0):
                make_squares(
                    0, 0, SCREENX, SCREENY, 1, "pink"
                )  # the border # fixme: pink
            else:
                make_squares(0, 0, SCREENX, SCREENY, 1, "#EEEEEE")  # the border

            # print the x and o at top
            # fixme:  data input could be corrupt, validate it x X o 0 O
            patterncount = 0
            for x in key:
                if x == "_":  # was failing if key had zero in position XXXXXX_10
                    break
                if x == "X":
                    make_labels(ALLX[patterncount] - 4, 75, 14, "X")
                if x == "0":
                    make_labels(ALLX[patterncount] - 4, 75, 14, "O")
                patterncount += 1

                # find the typo fixme:  single use to search my db for typo
                # if x =='x':
                #    print(name,key,position)

            for y in ALLY:  # draw our fret lines
                make_lines(ALLX[5], y, ALLX[0], y, 3)

            for x in ALLX:  # draw our string lines
                make_lines(x, ALLY[0], x, ALLY[4], 1)

            make_labels(261, 89, 16, position)  # which fret this pattern starts on
            offset = 330 - len(name) * 10 / 2  # a crude attempt at centering label
            make_labels(offset, 58, 18, name)  # the actual chordname at top

            # this is a work around because I loaded from sheet instead of json
            # converts [x,y,z],[x,y,z] into a true list from a string
            # fixme:  real data probably coming as json from sqlite
            fixedlist = fix_list(dots)

            if len(fixedlist) > 1:

                # however many dots we have, loop and place them
                for i in range(len(fixedlist)):

                    # chunk is 3 parts, we are moving around as needed to display them
                    # xyz[0] is x
                    # xyz[1] is y
                    # xyz[2] is the finger to use
                    chunk = fixedlist[i]
                    xyz = chunk.split(",")

                    xx = float(xyz[0])
                    yy = float(xyz[1])
                    text = 0  # give zz a default

                    if len(xyz) > 2:  # but if zz exists, then use it
                        text = xyz[2]

                    # fixme:  these do not scale, fix the offsets and radius to scale
                    make_dots(
                        xx, yy, 9, "white"
                    )  # this is the finger position, the actual dot
                    make_labels(
                        xx - 4, yy + 4, "14px", text
                    )  # this is the finger number inside the dot
            else:
                pass

            # write the file
            dwg.save()

    # requires "collections" module, sorts our dicitonary
    d_sorted = collections.OrderedDict(
        sorted(c_data.items(), key=lambda t: t[1]["name"])
    )

    create_html(d_sorted, ALLFILE)


def help_message(msg):
    """
    Shows help message/instructions
    """
    clear()
    print(Fore.RED + "-" * 55)
    print(Fore.RED + msg)
    print(Fore.RED + "-" * 55 + Fore.GREEN)
    # print(Style.BRIGHT + Fore.YELLOW + "\n(Argument Order is Specific!!)\n")
    print(Style.RESET_ALL + "Examples:")
    print(sys.argv[0] + " -h")
    print(sys.argv[0] + " -i <inputfile> -f -t somesheet")
    print(sys.argv[0] + " -i <inputfile> -p -t somesheet")
    print(sys.argv[0] + " -i <inputfile> -j -t somesheet")
    print(sys.argv[0] + " -i <inputfile> -c -t somesheet")
    print(sys.argv[0] + " -i <inputfile> -s -t somesheet")
    print(sys.argv[0] + " -i <inputfile> -l -t somesheet")
    print(sys.argv[0] + " -i <inputfile> -d <column> -t somesheet\n")

    print(Style.BRIGHT + Fore.YELLOW + "FLAG EXPLANATIONS\n")
    print("-h prints this help message")
    print("-i (INPUTFILE) inputfile is your source, input file, this is required")
    print("-f (FAVORITES) prints my favorite progressions")
    print("-p (PEEK) is to peek at the header and first row")
    print("-j (JSON) dumps json to screen, must come before -t")
    print("-c (ALL) svg files and loads them into chrome to view")
    print("-s (SCALES) creates scales files")
    print("-l (SHEETS) list sheet names")
    print("-d (DUMP) is to dump a column, you must specify which column")
    print(Style.RESET_ALL)
    list_files("xlsx")
    sys.exit(2)


def is_accessible(path, mode="r", encoding="utf-8"):
    """
    Check if the file or directory at `path` can
    be accessed by the program using `mode` open flags.
    Note: `encoding` is only used with text modes ('r', 'w', 'a', etc.).
    """
    try:
        # Only include encoding if mode indicates a text file
        if 'b' not in mode:
            f = open(path, mode, encoding=encoding)
        else:
            f = open(path, mode)
        f.close()
    except IOError:
        return False
    return True



def main():
    """
    checks for arguments, then educates if no/wrong arguments given
    checks if input file is readable, or error message
    """
    #workbook = ""  # fixme:  should I use a global or local to main?
    #sheetnames = ""  # fixme:  same as above
    #sheet = ""  # fixme:  same as above
    inputfile = ""  # fixme:  toxic input check
    #i, t = 0, 0

    # did they even give options?
    # if len(sys.argv) < 4:
    #    help_message("ADDITIONAL ARGUMENTS REQUIRED")

    # basic opt verification
    try:
        # anything with colon after requires argument
        # ? or h = helpmessage
        # t = "sheetname" is required for all options
        # p = requires i (input file).  Does a "peek" of your xlsx
        # j = requires i (input file).  Converts xlsx to json
        # c = requires i (input file).  Dumps entire chord database to html
        # f = requires i (input file).  Dumps my favorite progressions
        # s = requires i (input file).  Dumps scales
        # l = requires i (input file).  List Sheet Names
        # d = requires column # (input file too).  Dumps a specific column

        opts, args = getopt.getopt(sys.argv[1:], "h?pajcfsld:i:t:", ["help", "?"])

    except getopt.GetoptError as err:

        help_message("BAD OPTION")

    """
    maybe they typed weird option/arguments?
    maybe they asked for help?
    maybe they forgot to give an input file?
    maybe that file cannot be read
    Set the j and t flags properly
    """

    i = 0
    inputfile = ""

    # did they specify a file?
    for opt, arg in opts:

        if arg.startswith("-"):
            help_message(
                "ARGUMENT OMITTED FOR '{}'".format(opt) + " (ie:  -i somefile.xlsx)"
            )

        if opt in ("-h", "-help", "-?"):
            help_message("INSTRUCTIONS(1)")

        if opt in "-i":
            # verify exists and is readable
            if is_accessible(arg):
                i = 1
                inputfile = arg
            else:
                help_message("CANNOT READ FILE")

    if i:
        # cols = count_cols(sheetname)
        for opt, arg in opts:

            if opt in ("-a"):
                dump_all(inputfile)
    else:

        # WTF did they type?
        help_message("INSTRUCTIONS (MISSING FLAGS?)")

    sys.exit()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main()
    else:
        help_message("INSTRUCTIONS")
